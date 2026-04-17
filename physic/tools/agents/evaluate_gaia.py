import argparse
import json
import multiprocessing as mp
import os
import random
import re
import sys
import time
import traceback
import shutil
import warnings
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from dotenv import load_dotenv

    load_dotenv(encoding="utf-8")
except Exception:
    pass


def _ensure_physic_importable() -> None:
    here = Path(__file__).resolve()
    langchain_root = here.parents[3]
    if str(langchain_root) not in sys.path:
        sys.path.insert(0, str(langchain_root))


def _env_flag(key: str, default: bool = False) -> bool:
    raw = (os.environ.get(key) or "").strip().lower()
    if not raw:
        return bool(default)
    if raw in {"1", "true", "yes", "y", "on"}:
        return True
    if raw in {"0", "false", "no", "n", "off"}:
        return False
    return bool(default)


def _build_gaia_preamble(agent_tool_names: List[str], *, force_text_protocol: bool = False) -> str:
    tools = [t for t in agent_tool_names if t]
    tools_str = ", ".join(tools) if tools else "(none)"
    tool_calling_line = (
        "Use the text protocol for tool usage (one line). Do NOT use tool-calling JSON.\n"
        if force_text_protocol
        else "Never write tool calls as JSON text; invoke tools via the tool-calling mechanism.\n"
    )
    text_protocol_intro = "" if force_text_protocol else "If tool calling is unavailable, use this text protocol instead (one line):\n"
    output_rule = (
        "".join(
            [
                "Output exactly ONE line.\n",
                "If you need a tool, output ONLY the TOOL line (do NOT prefix it with 'Final answer:').\n",
                "Otherwise, output:\n",
                "Final answer: <answer>",
            ]
        )
        if force_text_protocol
        else "Output exactly ONE line, and ONLY the final answer, in this format:\nFinal answer: <answer>"
    )
    return "".join(
        [
            "You are an evaluation assistant for the GAIA benchmark.\n",
            f"Available tools: {tools_str}.\n",
            "Use the attachment preview/transcription/OCR as the primary source of truth.\n",
            "Only use web_search/read_webpage if the question requires external information not present in the attachment preview.\n",
            "Use python_calculator for arithmetic, counting, filtering, and simple parsing.\n",
            "When using python_calculator, prefer a single Python expression.\n",
            "If the attachment is CSV, you can assume a pandas DataFrame named df is available.\n",
            "If the attachment is JSON/JSONLD, you can assume a Python object named attachment is available.\n",
            "Do NOT output code, pseudocode, JSON tool-call blobs, or explanations.\n",
            tool_calling_line,
            text_protocol_intro,
            "- TOOL web_search: <query>\n",
            "- TOOL read_webpage: <url>\n",
            "- TOOL python_calculator: <python expression>\n",
            "When using the text protocol, output ONLY the TOOL line (do NOT prefix it with 'Final answer:').\n",
            "If no tools are available, you must still answer (do NOT output TOOL).\n",
            output_rule,
        ]
    )


def _extract_text_tool_call(text: str) -> Tuple[str, str]:
    t = (text or "").strip()
    if not t:
        return "", ""
    t2 = re.sub(r"(?im)^\s*Final\s*answer\s*:\s*", "", t).strip()
    t2 = re.sub(r"(?im)^\s*-\s*", "", t2).strip()
    patterns = [
        r"(?im)^\s*TOO[LL]\s+(web_search|read_webpage|python_calculator)\s*:\s*(.+?)\s*$",
        r"(?im)^\s*TOO[LL]\s*:\s*(web_search|read_webpage|python_calculator)\s*:\s*(.+?)\s*$",
        r"(?im)^\s*(web_search|read_webpage|python_calculator)\s*:\s*(.+?)\s*$",
    ]
    for p in patterns:
        matches = re.findall(p, t2)
        if matches:
            m = matches[-1]
            if len(m) == 2:
                name, arg = m[0], m[1]
            else:
                name, arg = m[0], m[1]
            return (str(name).strip(), str(arg).strip())
    inner = re.findall(r"(?im)\bTOO[LL]\s+(web_search|read_webpage|python_calculator)\s*:\s*(.+?)\s*$", t2)
    if inner:
        name, arg = inner[-1]
        return (str(name).strip(), str(arg).strip())
    return "", ""


def _web_search(query: str, max_results: int = 5) -> str:
    q = (query or "").strip()
    if not q:
        return "(empty query)"
    try:
        from duckduckgo_search import DDGS

        items = []
        with DDGS() as ddgs:
            for r in ddgs.text(q, max_results=max_results):
                title = (r.get("title") or "").strip()
                href = (r.get("href") or "").strip()
                body = (r.get("body") or "").strip()
                items.append(f"- {title}\n  {href}\n  {body}".strip())
        return "\n".join(items[:max_results]) if items else "(no results)"
    except Exception:
        try:
            import urllib.parse
            import urllib.request

            url = "https://html.duckduckgo.com/html/?q=" + urllib.parse.quote(q)
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                html = resp.read().decode("utf-8", errors="replace")
            try:
                from bs4 import BeautifulSoup

                soup = BeautifulSoup(html, "html.parser")
                out = []
                for a in soup.select("a.result__a")[:max_results]:
                    out.append(f"- {a.get_text(strip=True)}\n  {a.get('href')}")
                return "\n".join(out) if out else "(no results)"
            except Exception:
                return _clip_text(html, 4000)
        except Exception as e:
            return f"web_search failed: {e}"


def _read_webpage(url: str, max_chars: int = 8000) -> str:
    u = (url or "").strip()
    if not u:
        return "(empty url)"
    try:
        import urllib.request

        req = urllib.request.Request(u, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            html = resp.read().decode("utf-8", errors="replace")
        try:
            from bs4 import BeautifulSoup

            soup = BeautifulSoup(html, "html.parser")
            text = soup.get_text("\n", strip=True)
            return _clip_text(text, max_chars)
        except Exception:
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " ", text).strip()
            return _clip_text(text, max_chars)
    except Exception as e:
        return f"read_webpage failed: {e}"


def _maybe_prefetch_web_context(question: str, max_chars: int = 6000) -> str:
    if not _env_flag("GAIA_AUTO_PREFETCH_WEB", False):
        return ""
    q = (question or "").strip()
    if not q:
        return ""
    urls = re.findall(r"https?://[^\s\)\]\}>,]+", q)
    if urls:
        u = str(urls[0]).strip()
        return _read_webpage(u, max_chars=max_chars)
    if "wikipedia" in q.lower():
        sr = _web_search(q + " site:wikipedia.org", max_results=3)
        m = re.search(r"https?://[^\s\)\]\}>,]+", sr)
        if m:
            return _read_webpage(m.group(0), max_chars=max_chars)
    return ""


def _python_calculator(expr: str, locals_extra: Optional[Dict[str, Any]] = None) -> str:
    e = (expr or "").strip()
    if not e:
        return "(empty expression)"
    if (len(e) >= 2) and ((e[0] == e[-1] == "\"") or (e[0] == e[-1] == "'")):
        e = e[1:-1].strip()
    if "\n" in e:
        parts = [ln.strip() for ln in e.splitlines() if ln.strip()]
        if parts:
            e = parts[-1]
    if ";" in e:
        parts = [p.strip() for p in e.split(";") if p.strip()]
        if parts:
            e = parts[-1]
    if "import " in e:
        return "python_calculator failed: expression must be a single Python expression (no import)."
    import math
    import statistics
    import re

    safe_globals = {"__builtins__": {}}
    safe_locals = {
        "math": math,
        "statistics": statistics,
        "re": re,
        "round": round,
        "min": min,
        "max": max,
        "sum": sum,
        "len": len,
        "abs": abs,
        "float": float,
        "int": int,
        "str": str,
    }
    if locals_extra:
        safe_locals.update(locals_extra)
    try:
        val = eval(e, safe_globals, safe_locals)
    except Exception as ex:
        return f"python_calculator failed: {ex}"
    return str(val)


def _build_python_context_for_attachment(abs_attach: Optional[str], *, attachment_preview: Optional[str] = None) -> Dict[str, Any]:
    ctx: Dict[str, Any] = {}
    if not abs_attach:
        return ctx
    p = Path(str(abs_attach))
    ext = p.suffix.lower()
    ctx["attachment_path"] = str(p)
    if attachment_preview:
        ctx["attachment_text"] = str(attachment_preview)
    if ext in {".csv"}:
        try:
            import pandas as pd

            ctx["pd"] = pd
            ctx["df"] = pd.read_csv(str(p))
        except Exception:
            pass
    if ext in {".tsv"}:
        try:
            import pandas as pd

            ctx["pd"] = pd
            ctx["df"] = pd.read_csv(str(p), sep="\t")
        except Exception:
            pass
    if ext in {".txt"}:
        try:
            t = p.read_text(encoding="utf-8", errors="replace")
            ctx["attachment_text"] = t
        except Exception:
            pass
    if ext in {".xlsx", ".xls"}:
        try:
            import pandas as pd

            ctx["pd"] = pd
            df = pd.read_excel(str(p))
            ctx["df"] = df
        except Exception:
            pass
    if ext in {".json", ".jsonld"}:
        try:
            ctx["attachment"] = json.loads(p.read_text(encoding="utf-8", errors="replace"))
        except Exception:
            pass
    return ctx


def _run_one_query_subprocess(
    question: str,
    model_name: str,
    output_dir: str,
    verbose: bool,
    queue: "mp.Queue",
) -> None:
    try:
        _ensure_physic_importable()
        from physic.tools.agents.agent import CDEMAgentEvaluator

        defaults = _default_paths()
        evaluator = CDEMAgentEvaluator(
            vector_db_path=defaults["vector_db_path"],
            vector_db_collection=defaults["vector_db_collection"],
            test_data_dir=str(output_dir),
            output_dir=str(output_dir),
            model_name=model_name,
            enable_preprocessing=False,
            query_dataset_json=None,
        )
        out = evaluator.run_custom_query(query=question, case_filename=None, verbose=verbose)
        queue.put({"ok": True, "out": out})
    except Exception:
        queue.put({"ok": False, "error": traceback.format_exc()})


def _iter_jsonl(path: Path) -> Iterable[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = (line or "").strip()
            if not line:
                continue
            yield json.loads(line)


def _load_parquet_any(path: Path) -> List[Dict[str, Any]]:
    try:
        import pyarrow.parquet as pq
    except Exception as e:
        raise ImportError(f"缺少 pyarrow，无法读取 Parquet: {e}")
    table = pq.read_table(str(path))
    return table.to_pylist()


def _load_json_any(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    if path.suffix.lower() in {".parquet", ".pq"}:
        return _load_parquet_any(path)
    if path.suffix.lower() == ".jsonl":
        return list(_iter_jsonl(path))
    with open(path, "r", encoding="utf-8") as f:
        obj = json.load(f)
    if isinstance(obj, list):
        return [x for x in obj if isinstance(x, dict)]
    if isinstance(obj, dict):
        for k in ("data", "examples", "tasks", "items", "instances"):
            v = obj.get(k)
            if isinstance(v, list):
                return [x for x in v if isinstance(x, dict)]
        if "question" in obj or "answer" in obj:
            return [obj]
    raise ValueError(f"不支持的数据集格式: {path}")


def _pick_first(d: Dict[str, Any], keys: List[str]) -> Any:
    for k in keys:
        if k in d and d.get(k) is not None:
            return d.get(k)
    return None


def _strip_code_fences(text: str) -> str:
    t = (text or "").strip()
    if "```" not in t:
        return t
    pattern = r"```[\w+-]*\s*\n([\s\S]*?)```"
    matches = re.findall(pattern, t, re.IGNORECASE)
    if matches:
        inner = (matches[-1] or "").strip()
        return inner or t
    return t


def _remove_code_fences(text: str) -> str:
    t = (text or "")
    if "```" not in t:
        return t
    return re.sub(r"```[\w+-]*\s*\n[\s\S]*?```", "", t, flags=re.IGNORECASE)


def _extract_final_answer(text: str) -> str:
    t = (text or "").strip()
    if not t:
        return ""
    t_no_code = _remove_code_fences(t).strip()
    t2 = t_no_code if t_no_code else t
    patterns = [
        r"(?:final answer|final|answer)\s*[:：]\s*(.+)$",
        r"(?:答案|最终答案)\s*[:：]\s*(.+)$",
    ]
    for p in patterns:
        m = re.search(p, t2, flags=re.IGNORECASE | re.MULTILINE)
        if m:
            cand = (m.group(1) or "").strip()
            cand = cand.splitlines()[0].strip() if cand else ""
            if cand:
                return cand
    lines = [x.strip() for x in t2.splitlines() if x.strip()]
    if not lines:
        t3 = _strip_code_fences(t).strip()
        lines2 = [x.strip() for x in t3.splitlines() if x.strip()]
        return (lines2[-1] if lines2 else t3)[:200]
    last = lines[-1]
    if len(last) <= 200:
        return last
    if len(lines) >= 2 and len(lines[-2]) <= 200:
        return lines[-2]
    return last[:200]


def _normalize_answer(s: str) -> str:
    t = (s or "").strip().lower()
    if not t:
        return ""
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", t)
    t = re.sub(r"^the\s+|^a\s+|^an\s+", "", t)
    t = re.sub(r"\s+(the|a|an)\s+", " ", t)
    t = re.sub(r"[`\"“”‘’]", "", t)
    t = re.sub(r"[，,。！？!?\(\)\[\]\{\}<>]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _maybe_choice_letter(s: str) -> Optional[str]:
    t = _normalize_answer(s)
    if not t:
        return None
    if t in {"a", "b", "c", "d", "e"}:
        return t
    m = re.search(r"\b([a-e])\b", t)
    if m:
        return m.group(1)
    return None


def _parse_single_number(s: str) -> Optional[float]:
    t = (s or "").strip()
    if not t:
        return None
    t = t.replace(",", "")
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", t):
        return None
    try:
        return float(t)
    except Exception:
        return None


def _numeric_match(gold: str, pred: str, abs_tol: float = 1e-6, rel_tol: float = 1e-4) -> bool:
    g = _parse_single_number(gold)
    p = _parse_single_number(pred)
    if g is None or p is None:
        return False
    diff = abs(g - p)
    if diff <= abs_tol:
        return True
    denom = max(abs(g), abs(p), 1.0)
    return (diff / denom) <= rel_tol


@dataclass
class ScoredItem:
    item_id: str
    level: Optional[int]
    question: str
    gold: str
    pred_raw: str
    pred_final: str
    correct: bool
    match_type: str
    similarity: float


def _similarity(a: str, b: str) -> float:
    na = _normalize_answer(a)
    nb = _normalize_answer(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    import difflib
    return float(difflib.SequenceMatcher(a=na, b=nb).ratio())


def score_answer(gold: str, pred_raw: str) -> Tuple[bool, str, str, float]:
    gold_s = (gold or "").strip()
    pred_final = _extract_final_answer(pred_raw)
    if (not gold_s) or gold_s == "?":
        return False, "no_gold", pred_final, 0.0

    ng = _normalize_answer(gold_s)
    np = _normalize_answer(pred_final)
    sim = _similarity(ng, np)

    if ng in {"yes", "no"} and np in {"true", "false"}:
        np = "yes" if np == "true" else "no"
        sim = _similarity(ng, np)
    if ng in {"true", "false"} and np in {"yes", "no"}:
        ng = "yes" if ng == "true" else "no"
        sim = _similarity(ng, np)

    if ng and np and ng == np:
        return True, "exact", pred_final, sim

    gold_choice = _maybe_choice_letter(gold_s)
    if gold_choice:
        pred_choice = _maybe_choice_letter(pred_final)
        if pred_choice and pred_choice == gold_choice:
            return True, "choice", pred_final, sim

    if _numeric_match(gold_s, pred_final):
        return True, "numeric", pred_final, sim

    if ng and np and (ng in np):
        return True, "contains", pred_final, sim

    return False, "mismatch", pred_final, sim


def _default_paths() -> Dict[str, str]:
    here = Path(__file__).resolve()
    project_root = here.parents[2]
    repo_root = here.parents[3]

    def resolve_env_path(key: str, default_path: Path) -> str:
        raw = (os.environ.get(key) or "").strip()
        if not raw:
            return str(default_path)
        p = Path(raw)
        if not p.is_absolute():
            p = (repo_root / p).resolve()
        return str(p)

    default_vector_db_path = project_root / "tools" / "js_store" / "new_db_cdem"
    vector_db_path = resolve_env_path("CDEM_VECTOR_DB_PATH", default_vector_db_path)
    vector_db_collection = (os.environ.get("CDEM_VECTOR_DB_COLLECTION") or "new_db_cdem").strip()

    default_output_dir = here.parent / "results" / "gaia_evaluation_results"
    output_dir = resolve_env_path("GAIA_EVAL_OUTPUT_DIR", default_output_dir)

    return {
        "vector_db_path": vector_db_path,
        "vector_db_collection": vector_db_collection,
        "output_dir": output_dir,
    }


def _infer_gaia_repo_root(dataset_file: Path) -> Optional[Path]:
    p = dataset_file.resolve()
    if p.is_dir():
        return p
    parts = list(p.parts)
    for i, seg in enumerate(parts):
        if seg == "2023":
            if i == 0:
                return None
            return Path(*parts[:i])
    return p.parent


def _get_attachment_abs_path(item: Dict[str, Any], repo_root: Optional[Path], dataset_dir: Optional[Path]) -> Optional[str]:
    file_path = item.get("file_path")
    if file_path and str(file_path).strip():
        fp = str(file_path).strip()
        p = Path(fp)
        if p.is_absolute():
            return str(p.resolve())
        if repo_root is not None:
            return str((repo_root / p).resolve())
        if dataset_dir is not None:
            return str((dataset_dir / p).resolve())
        return fp

    file_name = item.get("file_name") or item.get("file") or item.get("attachment")
    if not file_name or not str(file_name).strip():
        return None
    fn = str(file_name).strip()
    p = Path(fn)
    if p.is_absolute():
        return str(p.resolve())
    if ("/" not in fn) and ("\\" not in fn) and dataset_dir is not None:
        return str((dataset_dir / fn).resolve())
    if repo_root is not None:
        return str((repo_root / p).resolve())
    if dataset_dir is not None:
        return str((dataset_dir / p).resolve())
    return fn


def _clip_text(s: str, max_chars: int) -> str:
    t = (s or "")
    if len(t) <= max_chars:
        return t
    return t[:max_chars] + "\n...（已截断）"


def _attachment_preview(abs_path: str, max_chars: int = 20000) -> str:
    p = Path((abs_path or "").strip())
    if not p.exists():
        return "Attachment missing."
    if p.is_dir():
        names = []
        for x in sorted(p.iterdir()):
            names.append(x.name + ("/" if x.is_dir() else ""))
            if len(names) >= 200:
                break
        return _clip_text("DIR:\n" + "\n".join(names), max_chars)

    ext = p.suffix.lower()
    if ext in {".txt", ".md", ".py", ".js", ".json", ".jsonld", ".xml", ".html", ".htm", ".pdb"}:
        data = p.read_bytes()
        text = ""
        for enc in ("utf-8", "utf-8-sig", "gb18030", "gbk", "cp936"):
            try:
                text = data.decode(enc)
                break
            except Exception:
                continue
        if not text:
            text = data.decode("utf-8", errors="replace")
        return _clip_text(text, max_chars)

    if ext in {".jsonl"}:
        lines = []
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            for i, line in enumerate(f, 1):
                if i > 80:
                    break
                lines.append(line.rstrip("\n"))
        return _clip_text("\n".join(lines), max_chars)

    if ext in {".csv", ".tsv"}:
        try:
            import pandas as pd
            df = pd.read_csv(str(p), nrows=50)
            return _clip_text(df.to_string(index=False), max_chars)
        except Exception:
            return _clip_text(p.read_text(encoding="utf-8", errors="replace"), max_chars)

    if ext in {".xlsx", ".xlsm", ".xls"}:
        try:
            import pandas as pd
            with pd.ExcelFile(str(p)) as xls:
                sheets = list(xls.sheet_names or [])
                out = [f"Excel sheets: {sheets}"]
                sheet0 = sheets[0] if sheets else None
            if sheet0:
                df = pd.read_excel(str(p), sheet_name=sheet0, nrows=30)
                out.append(f"\nSheet: {sheet0}")
                out.append(df.to_string(index=False))
                style_preview_enabled = (os.environ.get("GAIA_XLSX_STYLE_PREVIEW") or "").strip().lower() not in {"0", "false", "off", "no"}
                force_style_preview = (os.environ.get("GAIA_XLSX_STYLE_PREVIEW_FORCE") or "").strip().lower() in {"1", "true", "on", "yes"}
                try:
                    nan_ratio = float(df.isna().sum().sum()) / float(df.size or 1)
                except Exception:
                    nan_ratio = 0.0
                needs_style = force_style_preview or (nan_ratio >= 0.85) or bool(getattr(df, "empty", False)) or (len(getattr(df, "columns", [])) == 0)
                if style_preview_enabled and needs_style:
                    try:
                        import openpyxl

                        def _norm_rgb(v: str) -> str:
                            t = (v or "").strip().replace("#", "").upper()
                            if len(t) == 8:
                                t = t[2:]
                            return t if len(t) == 6 else ""

                        def _label_for(i: int) -> str:
                            alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                            if i < 26:
                                return alphabet[i]
                            a = i // 26 - 1
                            b = i % 26
                            return alphabet[a] + alphabet[b]

                        max_rows = int(os.environ.get("GAIA_XLSX_STYLE_MAX_ROWS") or "60")
                        max_cols = int(os.environ.get("GAIA_XLSX_STYLE_MAX_COLS") or "60")
                        max_rows = max(1, min(120, max_rows))
                        max_cols = max(1, min(120, max_cols))

                        wb = openpyxl.load_workbook(str(p), data_only=True)
                        try:
                            ws = wb.active
                            r_lim = min(int(getattr(ws, "max_row", 0) or 0), max_rows) or max_rows
                            c_lim = min(int(getattr(ws, "max_column", 0) or 0), max_cols) or max_cols

                            palette: dict[str, str] = {}
                            grid_lines: list[str] = []
                            for r in range(1, r_lim + 1):
                                row_tokens: list[str] = []
                                any_nonempty = False
                                for c in range(1, c_lim + 1):
                                    cell = ws.cell(row=r, column=c)
                                    v = cell.value
                                    if v is not None and str(v).strip() != "":
                                        s = str(v).strip().replace("\n", " ").replace("\r", " ")
                                        row_tokens.append(s[:6])
                                        any_nonempty = True
                                        continue

                                    rgb = ""
                                    try:
                                        fg = getattr(getattr(cell, "fill", None), "fgColor", None)
                                        if fg is not None and getattr(fg, "type", None) == "rgb":
                                            rgb = _norm_rgb(getattr(fg, "rgb", "") or "")
                                    except Exception:
                                        rgb = ""

                                    if rgb and rgb not in {"FFFFFF", "000000"}:
                                        if rgb not in palette:
                                            palette[rgb] = _label_for(len(palette))
                                        row_tokens.append(palette[rgb])
                                        any_nonempty = True
                                    else:
                                        row_tokens.append(".")

                                if any_nonempty and any(t != "." for t in row_tokens):
                                    grid_lines.append(" ".join(row_tokens).rstrip())

                            legend_lines = []
                            if palette:
                                legend_lines.append("Legend (fill colors):")
                                for rgb, lab in list(palette.items())[:80]:
                                    legend_lines.append(f"{lab}={rgb}")

                            style_txt = "Excel style preview (openpyxl fills):\n" + ("\n".join(grid_lines) if grid_lines else "(no styled cells detected)") + ("\n\n" + "\n".join(legend_lines) if legend_lines else "")
                            out.append("\n" + style_txt)
                        finally:
                            try:
                                wb.close()
                            except Exception:
                                pass
                    except Exception:
                        pass
            return _clip_text("\n".join(out), max_chars)
        except Exception:
            if ext == ".xls":
                return "Excel preview failed: .xls requires extra dependency (e.g. xlrd)."

            import zipfile
            import xml.etree.ElementTree as ET

            def _xlsx_preview_via_xml(pp: Path) -> str:
                with zipfile.ZipFile(str(pp), "r") as z:
                    names = set(z.namelist())
                    sheet_xml_name = None
                    if "xl/worksheets/sheet1.xml" in names:
                        sheet_xml_name = "xl/worksheets/sheet1.xml"
                    else:
                        for n in sorted(names):
                            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                                sheet_xml_name = n
                                break
                    if sheet_xml_name is None:
                        return "Excel preview failed: no worksheet xml found."

                    shared: list[str] = []
                    if "xl/sharedStrings.xml" in names:
                        try:
                            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
                            for node in root.iter():
                                if node.tag.endswith("}t") and node.text is not None:
                                    shared.append(node.text)
                        except Exception:
                            shared = []

                    def _cell_text(c_node) -> str:
                        t_attr = c_node.attrib.get("t")
                        v_node = None
                        for child in c_node:
                            if child.tag.endswith("}v"):
                                v_node = child
                                break
                            if child.tag.endswith("}is"):
                                for cc in child.iter():
                                    if cc.tag.endswith("}t") and cc.text is not None:
                                        return str(cc.text)
                        if v_node is None or v_node.text is None:
                            return ""
                        v = v_node.text
                        if t_attr == "s":
                            try:
                                idx = int(v)
                                return str(shared[idx]) if 0 <= idx < len(shared) else ""
                            except Exception:
                                return ""
                        return str(v)

                    sheet_root = ET.fromstring(z.read(sheet_xml_name))
                    cells: dict[tuple[int, int], str] = {}
                    max_r = 0
                    max_c = 0
                    for c in sheet_root.iter():
                        if not c.tag.endswith("}c"):
                            continue
                        ref = c.attrib.get("r") or ""
                        m = re.match(r"^([A-Z]+)(\d+)$", ref)
                        if not m:
                            continue
                        col_s, row_s = m.group(1), m.group(2)
                        try:
                            row_i = int(row_s)
                        except Exception:
                            continue
                        col_i = 0
                        for ch in col_s:
                            col_i = col_i * 26 + (ord(ch) - ord("A") + 1)
                        txt = _cell_text(c)
                        if txt == "":
                            continue
                        cells[(row_i, col_i)] = txt
                        max_r = max(max_r, row_i)
                        max_c = max(max_c, col_i)

                    max_rows = min(max_r or 50, 50)
                    max_cols = min(max_c or 20, 20)
                    lines = ["Excel preview (zip/xml fallback):"]
                    for r in range(1, max_rows + 1):
                        row_vals = []
                        for c in range(1, max_cols + 1):
                            v = (cells.get((r, c), "") or "").replace("\n", " ").replace("\r", " ").strip()
                            row_vals.append(v)
                        if any(x for x in row_vals):
                            lines.append("\t".join(row_vals).rstrip())
                    if len(lines) == 1:
                        lines.append("(no visible cells in preview window)")
                    return "\n".join(lines)

            try:
                txt = _xlsx_preview_via_xml(p)

                style_preview_enabled = (os.environ.get("GAIA_XLSX_STYLE_PREVIEW") or "").strip().lower() not in {"0", "false", "off", "no"}
                force_style_preview = (os.environ.get("GAIA_XLSX_STYLE_PREVIEW_FORCE") or "").strip().lower() in {"1", "true", "on", "yes"}

                needs_style = force_style_preview or ("(no visible cells in preview window)" in txt) or (txt.count("\n") <= 3)

                if style_preview_enabled and needs_style:
                    try:
                        import openpyxl
                        from openpyxl.utils import get_column_letter

                        def _norm_rgb(v: str) -> str:
                            t = (v or "").strip()
                            if not t:
                                return ""
                            t = t.replace("#", "").upper()
                            if len(t) == 8:
                                t = t[2:]
                            if len(t) != 6:
                                return ""
                            return t

                        def _label_for(i: int) -> str:
                            alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                            if i < 26:
                                return alphabet[i]
                            a = i // 26 - 1
                            b = i % 26
                            return alphabet[a] + alphabet[b]

                        max_rows = int(os.environ.get("GAIA_XLSX_STYLE_MAX_ROWS") or "60")
                        max_cols = int(os.environ.get("GAIA_XLSX_STYLE_MAX_COLS") or "60")
                        max_rows = max(1, min(120, max_rows))
                        max_cols = max(1, min(120, max_cols))

                        wb = openpyxl.load_workbook(str(p), data_only=True)
                        ws = wb.active
                        r_lim = min(int(getattr(ws, "max_row", 0) or 0), max_rows) or max_rows
                        c_lim = min(int(getattr(ws, "max_column", 0) or 0), max_cols) or max_cols

                        palette: dict[str, str] = {}
                        grid_lines: list[str] = []
                        for r in range(1, r_lim + 1):
                            row_tokens: list[str] = []
                            any_nonempty = False
                            for c in range(1, c_lim + 1):
                                cell = ws.cell(row=r, column=c)
                                v = cell.value
                                if v is not None and str(v).strip() != "":
                                    s = str(v).strip().replace("\n", " ").replace("\r", " ")
                                    tok = s[:6]
                                    row_tokens.append(tok)
                                    any_nonempty = True
                                    continue

                                rgb = ""
                                try:
                                    fill = cell.fill
                                    fg = getattr(fill, "fgColor", None)
                                    if fg is not None and getattr(fg, "type", None) == "rgb":
                                        rgb = _norm_rgb(getattr(fg, "rgb", "") or "")
                                except Exception:
                                    rgb = ""

                                if rgb and rgb not in {"FFFFFF", "000000"}:
                                    if rgb not in palette:
                                        palette[rgb] = _label_for(len(palette))
                                    row_tokens.append(palette[rgb])
                                    any_nonempty = True
                                else:
                                    row_tokens.append(".")

                            if any_nonempty and any(t != "." for t in row_tokens):
                                grid_lines.append(" ".join(row_tokens).rstrip())

                        legend_lines = []
                        if palette:
                            legend_lines.append("Legend (fill colors):")
                            for rgb, lab in list(palette.items())[:80]:
                                legend_lines.append(f"{lab}={rgb}")

                        style_txt = "Excel style preview (openpyxl fills):\n" + ("\n".join(grid_lines) if grid_lines else "(no styled cells detected)") + ("\n\n" + "\n".join(legend_lines) if legend_lines else "")
                        txt = txt + "\n\n" + style_txt
                    except Exception:
                        pass

                return _clip_text(txt, max_chars)
            except Exception as e:
                return f"Excel preview failed: {e}"

    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(str(p))
            out = [f"PDF pages: {doc.page_count}"]
            for i in range(min(6, doc.page_count)):
                txt = (doc.load_page(i).get_text("text") or "").strip()
                if txt:
                    out.append(f"\n--- page {i+1} ---\n{txt}")
            doc.close()
            return _clip_text("\n".join(out), max_chars)
        except Exception:
            try:
                from pypdf import PdfReader
                reader = PdfReader(str(p))
                out = [f"PDF pages: {len(reader.pages)}"]
                for i in range(min(6, len(reader.pages))):
                    page = reader.pages[i]
                    txt = (page.extract_text() or "").strip()
                    if txt:
                        out.append(f"\n--- page {i+1} ---\n{txt}")
                return _clip_text("\n".join(out), max_chars)
            except Exception:
                import shutil
                import subprocess
                exe = shutil.which("pdftotext")
                if not exe:
                    return "PDF preview failed: missing dependency (install pymupdf/fitz, or pypdf, or system pdftotext)."
                try:
                    res = subprocess.run(
                        [exe, "-f", "1", "-l", "6", "-layout", "-nopgbrk", str(p), "-"],
                        capture_output=True,
                        text=True,
                        timeout=20,
                    )
                    txt = (res.stdout or "").strip()
                    if not txt:
                        return "PDF preview failed: pdftotext produced empty output."
                    return _clip_text("PDF preview (pdftotext):\n" + txt, max_chars)
                except Exception as e:
                    return f"PDF preview failed: {e}"

    if ext in {".docx", ".pptx", ".zip"}:
        import zipfile
        import xml.etree.ElementTree as ET
        try:
            with zipfile.ZipFile(str(p), "r") as z:
                names = z.namelist()
                if ext == ".docx" and "word/document.xml" in names:
                    xml_bytes = z.read("word/document.xml")
                    root = ET.fromstring(xml_bytes)
                    lines: list[str] = []
                    for para in root.iter():
                        if not para.tag.endswith("}p"):
                            continue
                        parts: list[str] = []
                        for node in para.iter():
                            if node.tag.endswith("}t") and node.text:
                                parts.append(node.text)
                        if not parts:
                            continue
                        s = " ".join([x.strip() for x in parts if x.strip()])
                        s = re.sub(r"\s+", " ", s).strip()
                        if s:
                            lines.append(s)
                    if not lines:
                        texts = []
                        for node in root.iter():
                            if node.tag.endswith("}t") and node.text:
                                texts.append(node.text)
                        return _clip_text(" ".join([x.strip() for x in texts if x and x.strip()]), max_chars)
                    return _clip_text("\n".join(lines), max_chars)
                if ext == ".pptx":
                    slide_names = sorted([n for n in names if n.startswith("ppt/slides/slide") and n.endswith(".xml")])
                    chunks = []
                    for n in slide_names[:20]:
                        xml_bytes = z.read(n)
                        try:
                            root = ET.fromstring(xml_bytes)
                        except Exception:
                            continue
                        texts = []
                        for node in root.iter():
                            if node.tag.endswith("}t") and node.text:
                                texts.append(node.text)
                        if texts:
                            s = " ".join([x.strip() for x in texts if x and x.strip()])
                            s = re.sub(r"\s+", " ", s).strip()
                            chunks.append(f"[{Path(n).name}]\n" + s)
                    if chunks:
                        return _clip_text("\n\n".join(chunks), max_chars)
                if ext == ".zip":
                    import tempfile
                    out = ["ZIP entries:\n" + "\n".join(names[:200])]
                    previewed = 0
                    for n in names:
                        if previewed >= 3:
                            break
                        low = n.lower()
                        if not any(low.endswith(s) for s in (".txt", ".md", ".csv", ".tsv", ".json", ".jsonld", ".xml", ".html", ".htm", ".pdb", ".py", ".js", ".xlsx", ".xlsm", ".xls", ".pdf")):
                            continue
                        try:
                            data = z.read(n)
                        except Exception:
                            continue
                        try:
                            with tempfile.NamedTemporaryFile(delete=True, suffix=Path(n).suffix) as tf:
                                tf.write(data)
                                tf.flush()
                                out.append(f"\n\n--- {n} ---\n" + _attachment_preview(tf.name, max_chars=max(2000, min(max_chars, 12000))))
                            previewed += 1
                        except Exception:
                            continue
                    return _clip_text("\n".join(out), max_chars)
                return _clip_text("ZIP entries:\n" + "\n".join(names[:200]), max_chars)
        except Exception as e:
            return f"Archive preview failed: {e}"

    if ext in {".png", ".jpg", ".jpeg", ".webp"}:
        enable_ocr = (os.environ.get("GAIA_ENABLE_OCR") or "").strip().lower() in {"1", "true", "on", "yes"}
        try:
            from PIL import Image
            im = Image.open(str(p))
            header = f"Image: {im.format} {im.size[0]}x{im.size[1]} mode={im.mode}"
            if not enable_ocr:
                return header + "\nOCR not enabled."
        except Exception as e:
            if not enable_ocr:
                return f"Image preview failed: {e}"
            header = "Image OCR enabled."

        import shutil
        import subprocess
        exe = (os.environ.get("TESSERACT_CMD") or "").strip() or shutil.which("tesseract")
        if not exe:
            return header + "\nOCR failed: system tesseract not found."
        lang = (os.environ.get("GAIA_TESSERACT_LANG") or "eng").strip() or "eng"
        try:
            res = subprocess.run([exe, str(p), "stdout", "-l", lang], capture_output=True, text=True, timeout=60)
            txt = (res.stdout or "").strip()
            if not txt:
                return header + "\nImage OCR: (empty)"
            return _clip_text(header + "\nImage OCR:\n" + txt, max_chars)
        except Exception as ee:
            return header + f"\nOCR failed: {ee}"

    if ext in {".mp3", ".m4a", ".wav", ".mov"}:
        enable_asr = (os.environ.get("GAIA_ENABLE_ASR") or "").strip().lower() in {"1", "true", "on", "yes"}
        if not enable_asr:
            return "Binary media file (audio/video). Offline transcription not enabled."
        import importlib.util
        import shutil
        import subprocess

        exe = shutil.which("whisper")
        model = (os.environ.get("GAIA_WHISPER_MODEL") or "base").strip() or "base"

        runner: Optional[List[str]] = None
        if exe:
            runner = [exe]
        elif importlib.util.find_spec("whisper") is not None:
            runner = [sys.executable, "-m", "whisper"]
        else:
            return "Binary media file (audio/video). Transcription enabled but whisper not found."

        ffmpeg = (os.environ.get("FFMPEG_CMD") or "").strip() or shutil.which("ffmpeg")
        if not ffmpeg:
            return "Binary media file (audio/video). Transcription enabled but ffmpeg not found."
        try:
            import tempfile
            with tempfile.TemporaryDirectory() as td:
                env = os.environ.copy()
                if ffmpeg and (os.path.isabs(ffmpeg) or ("\\" in ffmpeg) or ("/" in ffmpeg)):
                    ff_dir = str(Path(ffmpeg).resolve().parent)
                    env["PATH"] = ff_dir + os.pathsep + env.get("PATH", "")
                res = subprocess.run(
                    [*runner, str(p), "--model", model, "--output_format", "txt", "--output_dir", td, "--fp16", "False"],
                    capture_output=True,
                    text=True,
                    env=env,
                    timeout=240,
                )
                txt_files = list(Path(td).glob("*.txt"))
                if not txt_files:
                    return "Audio transcription failed: no txt output."
                text = txt_files[0].read_text(encoding="utf-8", errors="replace").strip()
                if not text:
                    return "Audio transcription: (empty)"
                return _clip_text("Audio transcription:\n" + text, max_chars)
        except Exception as e:
            return f"Audio transcription failed: {e}"

    return f"Unsupported attachment type: {ext}"


def _download_gaia_hf(
    local_dir: Path,
    subset: str = "2023_all",
    split: str = "validation",
    endpoint: Optional[str] = None,
    token: Optional[str] = None,
) -> Path:
    try:
        from huggingface_hub import snapshot_download
    except Exception as e:
        raise ImportError(f"缺少 huggingface_hub: {e}")

    local_dir = local_dir.resolve()
    local_dir.mkdir(parents=True, exist_ok=True)

    allow_patterns = [
        "README.md",
        ".gitattributes",
        f"2023/{split}/*",
        f"2023/{split}/**",
    ]

    endpoint = (endpoint or "").strip() or None
    if endpoint is None:
        endpoint = (os.environ.get("HF_ENDPOINT") or "").strip() or None
    if endpoint is None:
        endpoint = (os.environ.get("HUGGINGFACE_HUB_BASE_URL") or "").strip() or None
    token = (token or "").strip() or None
    if token is None:
        token = (os.environ.get("HF_TOKEN") or "").strip() or None
    if token is None:
        token = (os.environ.get("HUGGINGFACE_HUB_TOKEN") or "").strip() or None

    data_dir = Path(
        snapshot_download(
            repo_id="gaia-benchmark/GAIA",
            repo_type="dataset",
            local_dir=str(local_dir),
            allow_patterns=allow_patterns,
            endpoint=endpoint,
            token=token,
        )
    )

    subset = (subset or "").strip()
    split = (split or "").strip()
    if not subset or subset == "2023_all":
        parquet_rel = f"2023/{split}/metadata.parquet"
    elif subset in {"2023_level1", "2023_level2", "2023_level3"}:
        lvl = subset.split("_")[-1].replace("level", "").strip()
        parquet_rel = f"2023/{split}/metadata.level{lvl}.parquet"
    else:
        raise ValueError(f"未知 subset: {subset}（可选 2023_all/2023_level1/2023_level2/2023_level3）")

    parquet_path = data_dir / parquet_rel
    if not parquet_path.exists():
        raise FileNotFoundError(f"未找到 GAIA metadata parquet: {parquet_path}")
    return parquet_path


def evaluate_gaia(
    dataset_path: str,
    model_name: str,
    output_dir: str,
    backend: str = "cdem",
    limit: Optional[int] = None,
    offset: int = 0,
    shuffle: bool = False,
    seed: int = 42,
    level: Optional[int] = None,
    only_with_file: bool = False,
    inline_attachment: bool = True,
    inline_max_chars: int = 20000,
    question_key: Optional[str] = None,
    answer_key: Optional[str] = None,
    id_key: Optional[str] = None,
    verbose: bool = False,
    dry_run: bool = False,
    timeout_s: Optional[float] = None,
    gaia_preamble: bool = True,
    preamble_text: Optional[str] = None,
    format_retry: bool = True,
    format_retry_rounds: int = 1,
) -> Dict[str, Any]:
    ds_path = Path(dataset_path)
    items = _load_json_any(ds_path)
    if level is not None:
        want = int(level)
        filtered: List[Dict[str, Any]] = []
        for it in items:
            try:
                if int(it.get("Level")) == want:
                    filtered.append(it)
            except Exception:
                continue
        items = filtered
    if only_with_file:
        filtered2: List[Dict[str, Any]] = []
        for it in items:
            fn = it.get("file_name")
            if fn and str(fn).strip():
                filtered2.append(it)
        items = filtered2
    if shuffle:
        rng = random.Random(seed)
        rng.shuffle(items)
    if offset:
        items = items[offset:]
    if limit is not None:
        items = items[: max(0, int(limit))]
    total = len(items)
    if total == 0:
        return {"ok": False, "error": "数据集任务列表为空"}

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    pred_path = out_dir / f"predictions_{run_id}.jsonl"
    submission_path = out_dir / f"submission_{run_id}.jsonl"
    summary_path = out_dir / f"summary_{run_id}.json"

    evaluator = None
    agent_tool_names: List[str] = []
    force_text_protocol = False
    if not dry_run:
        _ensure_physic_importable()
        backend_s = (backend or "").strip().lower() or "cdem"
        force_text_protocol = backend_s in {"evo", "evolutionary"} or backend_s.startswith("metagpt")
        if backend_s in {"metagpt", "metagpt_evo"}:
            from physic.tools.agents.metagpt_evaluator import MetaGPTEvaluator

            use_evo = backend_s == "metagpt_evo" or _as_bool(os.environ.get("METAGPT_USE_EVO"), default=False)
            evaluator = MetaGPTEvaluator(model_name=model_name, use_evo=use_evo)
            agent_tool_names = []
        elif backend_s in {"evo", "evolutionary"}:
            from physic.tools.agents.agent import CDEMAgentEvaluator
            from physic.tools.agents.evolutionary_single_agent_langgraph import EvolutionaryWrapperAgentConstructionModule

            defaults = _default_paths()
            evaluator = CDEMAgentEvaluator(
                vector_db_path=defaults["vector_db_path"],
                vector_db_collection=defaults["vector_db_collection"],
                test_data_dir=str(out_dir),
                output_dir=str(out_dir),
                model_name=model_name,
                enable_preprocessing=False,
                query_dataset_json=None,
            )
            try:
                tools_obj = getattr(getattr(evaluator, "agent_module", None), "tools", None) or []
                for t in tools_obj:
                    name = getattr(t, "name", None)
                    if name:
                        agent_tool_names.append(str(name))
            except Exception:
                agent_tool_names = []
            wrapper = EvolutionaryWrapperAgentConstructionModule(base_agent=evaluator.agent_module, domain="gaia_text")
            evaluator.agent_module = wrapper
            evaluator.eval_module.agent_module = wrapper
        else:
            from physic.tools.agents.agent import CDEMAgentEvaluator

            defaults = _default_paths()
            evaluator = CDEMAgentEvaluator(
                vector_db_path=defaults["vector_db_path"],
                vector_db_collection=defaults["vector_db_collection"],
                test_data_dir=str(out_dir),
                output_dir=str(out_dir),
                model_name=model_name,
                enable_preprocessing=False,
                query_dataset_json=None,
            )
            try:
                tools_obj = getattr(getattr(evaluator, "agent_module", None), "tools", None) or []
                for t in tools_obj:
                    name = getattr(t, "name", None)
                    if name:
                        agent_tool_names.append(str(name))
            except Exception:
                agent_tool_names = []

        has_file_tools = any(x in {"read_local_file", "list_local_dir"} for x in agent_tool_names)
        if (not bool(inline_attachment)) and (not has_file_tools):
            inline_attachment = True

    correct = 0
    scored: List[ScoredItem] = []
    started = time.time()

    q_keys = [question_key] if question_key else []
    q_keys += ["Question", "question", "query", "prompt", "instruction", "input", "problem"]
    a_keys = [answer_key] if answer_key else []
    a_keys += ["Final answer", "answer", "final_answer", "gold", "label", "output", "target"]
    i_keys = [id_key] if id_key else []
    i_keys += ["task_id", "id", "uid", "example_id", "name"]

    repo_root = _infer_gaia_repo_root(ds_path)
    dataset_dir = ds_path.resolve().parent
    gaia_preamble_text = (preamble_text or os.environ.get("GAIA_PROMPT_PREAMBLE") or "").strip()
    if gaia_preamble and not gaia_preamble_text:
        gaia_preamble_text = _build_gaia_preamble(agent_tool_names, force_text_protocol=force_text_protocol)

    with open(pred_path, "w", encoding="utf-8") as f, open(submission_path, "w", encoding="utf-8") as sf:
        for idx, item in enumerate(items, 1):
            item_id = str(_pick_first(item, i_keys) or f"I{idx:06d}")
            question_raw = str(_pick_first(item, q_keys) or "").strip()
            try:
                lvl_val = int(item.get("Level")) if item.get("Level") is not None else None
            except Exception:
                lvl_val = None
            abs_attach = _get_attachment_abs_path(item=item, repo_root=repo_root, dataset_dir=dataset_dir)
            question = (question_raw or "").strip()
            if gaia_preamble and gaia_preamble_text:
                question = gaia_preamble_text + "\n\n" + question
            preview: str = ""
            if abs_attach:
                question += "\n\n" + f"Attachment path: {abs_attach}"
                if inline_attachment:
                    eff_max_chars = int(inline_max_chars)
                    if lvl_val == 3:
                        try:
                            eff_max_chars = max(eff_max_chars, int(os.environ.get("GAIA_LEVEL3_INLINE_MAX_CHARS", "60000")))
                        except Exception:
                            eff_max_chars = max(eff_max_chars, 60000)
                    preview = _attachment_preview(abs_attach, max_chars=int(eff_max_chars)) or ""
                    if preview:
                        question += "\n\nAttachment preview:\n" + preview
            else:
                prefetch = _maybe_prefetch_web_context(question_raw)
                if prefetch:
                    question += "\n\nPrefetched webpage context:\n" + prefetch
            gold = str(_pick_first(item, a_keys) or "").strip()
            if not question:
                continue

            pred_raw = ""
            gen_meta: Dict[str, Any] = {}
            if evaluator is not None:
                try:
                    if timeout_s and float(timeout_s) > 0:
                        q: "mp.Queue" = mp.Queue()
                        p = mp.Process(
                            target=_run_one_query_subprocess,
                            args=(question, model_name, str(out_dir), bool(verbose), q),
                        )
                        p.start()
                        p.join(float(timeout_s))
                        if p.is_alive():
                            p.terminate()
                            p.join(1)
                            out = {"generated_code": "", "generation_time": float(timeout_s), "retrieved_docs_count": 0, "task_steps": [], "last_run_metrics": {}}
                            gen_meta = {"error": "timeout"}
                        else:
                            try:
                                child = q.get_nowait()
                            except Exception:
                                child = {"ok": False, "error": "no_return"}
                            if child.get("ok"):
                                out = dict(child.get("out") or {})
                            else:
                                out = {"generated_code": "", "generation_time": 0.0, "retrieved_docs_count": 0, "task_steps": [], "last_run_metrics": {}}
                                gen_meta = {"error": str(child.get("error") or "")[:4000]}
                    else:
                        out = evaluator.run_custom_query(question, case_filename=None, verbose=verbose)

                    pred_raw = str(out.get("generated_code") or "")
                    gen_meta = {
                        "generation_time": float(out.get("generation_time") or 0.0),
                        "retrieved_docs_count": int(out.get("retrieved_docs_count") or 0),
                        "task_steps": list(out.get("task_steps") or []),
                        "last_run_metrics": dict(out.get("last_run_metrics") or {}),
                        **gen_meta,
                    }
                except Exception as e:
                    pred_raw = ""
                    gen_meta = {"error": str(e)}

            if evaluator is not None:
                seen_tool_calls: set = set()
                max_tool_steps = 6
                if lvl_val == 3:
                    try:
                        max_tool_steps = max(max_tool_steps, int(os.environ.get("GAIA_LEVEL3_MAX_TOOL_STEPS", "10")))
                    except Exception:
                        max_tool_steps = max(max_tool_steps, 10)

                if lvl_val == 3 and abs_attach:
                    tn0, _ = _extract_text_tool_call(pred_raw)
                    if not tn0:
                        try:
                            nudge = (
                                "This is a Level 3 GAIA question and includes an attachment.\n"
                                "Before answering, you MUST use python_calculator to parse/extract/compute from the attachment preview.\n"
                                "Output ONLY one line in the text protocol:\n"
                                "TOOL python_calculator: <one Python expression>\n\n"
                                f"{question}\n"
                            )
                            out_n = evaluator.run_custom_query(nudge, case_filename=None, verbose=False)
                            pred_raw_n = str(out_n.get("generated_code") or "").strip()
                            if pred_raw_n:
                                pred_raw = pred_raw_n
                                gen_meta = dict(gen_meta)
                                gen_meta["gaia_level3_tool_nudge"] = True
                        except Exception:
                            pass

                for _ in range(int(max_tool_steps)):
                    tool_name, tool_arg = _extract_text_tool_call(pred_raw)
                    if not tool_name:
                        break
                    sig = (tool_name, tool_arg)
                    if sig in seen_tool_calls:
                        break
                    seen_tool_calls.add(sig)
                    try:
                        if tool_name == "web_search":
                            tool_out = _web_search(tool_arg)
                        elif tool_name == "read_webpage":
                            tool_out = _read_webpage(tool_arg)
                        elif tool_name == "python_calculator":
                            ctx = _build_python_context_for_attachment(abs_attach, attachment_preview=preview)
                            tool_out = _python_calculator(tool_arg, locals_extra=ctx)
                        else:
                            break
                        if tool_name == "web_search":
                            try:
                                urls = re.findall(r"https?://\\S+", str(tool_out))
                            except Exception:
                                urls = []
                            if urls:
                                top_url = urls[0].rstrip(").,;\"'")
                                try:
                                    page = _read_webpage(top_url)
                                    tool_out = (str(tool_out).rstrip() + "\n\nAuto-read top result:\n" + str(page).strip()).strip()
                                except Exception:
                                    pass
                        if tool_name == "python_calculator" and str(tool_out).startswith("python_calculator failed:"):
                            fix_followup = (
                                "Your python_calculator expression failed.\n\n"
                                f"Error: {tool_out}\n\n"
                                "Output ONLY one line in the text protocol with a corrected single Python expression:\n"
                                "TOOL python_calculator: <python expression>\n\n"
                                "Notes:\n"
                                "- If a DataFrame is available, use df.\n"
                                "- If a JSON object is available, use attachment.\n"
                                "- You can use attachment_text (string) for regex extraction.\n\n"
                                f"{question}\n"
                            )
                            out_fix = evaluator.run_custom_query(fix_followup, case_filename=None, verbose=False)
                            pred_raw = str(out_fix.get("generated_code") or "").strip()
                            gen_meta = dict(gen_meta)
                            gen_meta["text_tool_protocol"] = True
                            gen_meta["python_calculator_fixup"] = True
                            continue
                        followup = (
                            "You requested a tool. Here is the tool result.\n\n"
                            f"Tool: {tool_name}\n"
                            f"Input: {tool_arg}\n"
                            "Output:\n"
                            f"{tool_out}\n\n"
                            "Now answer the GAIA question.\n"
                            "Output exactly ONE line:\n"
                            "Final answer: <answer>\n\n"
                            f"{question}\n"
                        )
                        out_t = evaluator.run_custom_query(followup, case_filename=None, verbose=False)
                        pred_raw = str(out_t.get("generated_code") or "").strip()
                        gen_meta = dict(gen_meta)
                        gen_meta["text_tool_protocol"] = True
                    except Exception:
                        break

            if evaluator is not None and bool(format_retry) and int(format_retry_rounds or 0) > 0:
                pred_final_probe = _extract_final_answer(pred_raw)

                def _bad_final(ans: str, raw: str) -> bool:
                    a = (ans or "").strip()
                    r = (raw or "").strip()
                    if not a:
                        return True
                    if a.startswith("{") and a.endswith("}"):
                        return True
                    if re.search(r"(?i)\bTOO[LL]\b\s+(web_search|read_webpage|python_calculator)\s*:", a):
                        return True
                    if "资料不足" in a or "insufficient" in a.lower():
                        return True
                    if "```" in r:
                        return True
                    if "\"name\"" in r and ("python_calculator" in r or "web_search" in r or "read_webpage" in r):
                        return True
                    if ("{" in a) or ("}" in a):
                        return True
                    return False

                if _bad_final(pred_final_probe, pred_raw):
                    for _ in range(max(1, int(format_retry_rounds))):
                        try:
                            fix_prompt = (
                                "Rewrite the final answer only.\n"
                                "Output exactly ONE line:\n"
                                "Final answer: <answer>\n"
                                "Do NOT output code, JSON, or any extra text.\n\n"
                                "GAIA question (with attachment preview):\n"
                                f"{question}\n\n"
                                "Previous response:\n"
                                f"{pred_raw}\n"
                            )
                            out2 = evaluator.run_custom_query(fix_prompt, case_filename=None, verbose=False)
                            pred_raw2 = str(out2.get("generated_code") or "").strip()
                            if pred_raw2:
                                pred_raw = pred_raw2
                                gen_meta = dict(gen_meta)
                                gen_meta["format_retry"] = True
                                break
                        except Exception:
                            break

            is_ok, match_type, pred_final, sim = score_answer(gold=gold, pred_raw=pred_raw)
            if is_ok:
                correct += 1
            scored_item = ScoredItem(
                item_id=item_id,
                level=lvl_val,
                question=question,
                gold=gold,
                pred_raw=pred_raw,
                pred_final=pred_final,
                correct=bool(is_ok),
                match_type=str(match_type),
                similarity=float(sim),
            )
            scored.append(scored_item)
            rec = asdict(scored_item)
            rec["meta"] = gen_meta
            rec["raw_item"] = item
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

            sf.write(
                json.dumps(
                    {
                        "task_id": item_id,
                        "Final answer": (pred_final or "").strip(),
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )

    elapsed = time.time() - started
    denom = max(len(scored), 1)
    accuracy = correct / denom

    by_type: Dict[str, int] = {}
    for s in scored:
        by_type[s.match_type] = by_type.get(s.match_type, 0) + 1

    by_level: Dict[str, Dict[str, float]] = {}
    by_level_named: Dict[str, Dict[str, float]] = {}
    lvl_counts: Dict[int, int] = {}
    lvl_correct: Dict[int, int] = {}
    for s in scored:
        try:
            lvl = int(s.level) if s.level is not None else -1
        except Exception:
            lvl = -1
        if lvl in {1, 2, 3}:
            lvl_counts[lvl] = lvl_counts.get(lvl, 0) + 1
            if bool(s.correct):
                lvl_correct[lvl] = lvl_correct.get(lvl, 0) + 1
    for lvl in (1, 2, 3):
        total_l = int(lvl_counts.get(lvl) or 0)
        correct_l = int(lvl_correct.get(lvl) or 0)
        rec = {
            "total": total_l,
            "correct": correct_l,
            "accuracy": float(round((correct_l / max(total_l, 1)), 6)) if total_l > 0 else 0.0,
        }
        by_level[str(lvl)] = rec
        by_level_named[f"level{lvl}"] = rec

    summary = {
        "ok": True,
        "dataset": str(ds_path),
        "model": model_name,
        "backend": str(backend or "cdem"),
        "dry_run": bool(dry_run),
        "total": int(len(scored)),
        "correct": int(correct),
        "accuracy": float(round(accuracy, 6)),
        "by_level": by_level,
        "by_level_named": by_level_named,
        "match_type_counts": by_type,
        "output_predictions_jsonl": str(pred_path),
        "output_submission_jsonl": str(submission_path),
        "elapsed_s": float(round(elapsed, 3)),
        "agent_tools": agent_tool_names,
        "effective_inline_attachment": bool(inline_attachment),
    }
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return summary


def _run_self_test() -> int:
    tests = [
        ("42", "Final answer: 42", True, "exact"),
        ("3.14", "答案：3.1400000", True, "numeric"),
        ("A", "I think the answer is (A).", True, "choice"),
        ("New York", "Final answer: New York City", True, "contains"),
        ("foo", "bar", False, "mismatch"),
    ]
    ok = 0
    for gold, pred, exp_ok, exp_type in tests:
        got_ok, got_type, _, _ = score_answer(gold, pred)
        if bool(got_ok) == bool(exp_ok) and str(got_type) == str(exp_type):
            ok += 1
    return 0 if ok == len(tests) else 2


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset", type=str, default=os.environ.get("GAIA_DATASET_PATH") or "", help="GAIA 数据集文件路径（.json/.jsonl/.parquet）")
    ap.add_argument("--model", type=str, default=(os.environ.get("CDEM_LLM_MODEL") or "qwen3.5-flash"), help="模型名")
    ap.add_argument("--backend", type=str, default=(os.environ.get("GAIA_BACKEND") or "cdem"), help="cdem/metagpt")
    ap.add_argument("--output_dir", type=str, default=(_default_paths()["output_dir"]), help="输出目录")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--offset", type=int, default=0)
    ap.add_argument("--shuffle", action="store_true")
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--level", type=int, default=None, help="仅评测某个难度等级（1/2/3），默认不过滤")
    ap.add_argument("--only_with_file", action="store_true", help="只评测带附件的题（file_name 非空）")
    ap.add_argument("--inline_attachment", action="store_true", default=True, help="将附件内容预览直接拼进提示词（离线解题推荐；默认开启）")
    ap.add_argument("--no_inline_attachment", action="store_true")
    ap.add_argument("--inline_max_chars", type=int, default=20000)
    ap.add_argument("--force_inline_attachment", action="store_true", help="强制 inline 附件预览（即使传了 --no_inline_attachment 也会开启；用于没有文件工具时的 GAIA 评测）")
    ap.add_argument("--question_key", type=str, default=None)
    ap.add_argument("--answer_key", type=str, default=None)
    ap.add_argument("--id_key", type=str, default=None)
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--dry_run", action="store_true", help="不调用智能体，仅输出空预测并进行打分（用于验证加载/字段）")
    ap.add_argument("--timeout_s", type=float, default=None, help="单题超时（秒）。设置后每题在子进程中运行以便超时中断，会显著变慢。")
    ap.add_argument("--no_force_gaia_mode", action="store_true", help="不强制设置 GAIA_MODE=1（默认会强制，以避免被 .env 覆盖导致走非 GAIA prompt/工具链）")
    ap.add_argument("--no_gaia_preamble", action="store_true", help="不在问题前添加 GAIA 通用前导指令（默认添加，以提升工具调用与输出格式稳定性）")
    ap.add_argument("--gaia_preamble_text", type=str, default="", help="自定义 GAIA 前导指令（覆盖默认/环境变量 GAIA_PROMPT_PREAMBLE）")
    ap.add_argument("--gaia_preamble_file", type=str, default="", help="从文件读取 GAIA 前导指令（优先级高于 --gaia_preamble_text 与环境变量）")
    ap.add_argument("--no_force_task_type", action="store_true", help="不强制将 CDEM_TASK_TYPE 设为 qa（默认 GAIA 评测会强制，以避免误走脚本生成分支）")
    ap.add_argument("--no_format_retry", action="store_true", help="不做格式化二次重试（默认开启：当输出为代码/JSON/资料不足/无Final时会再追问一次只输出Final answer）")
    ap.add_argument("--format_retry_rounds", type=int, default=1, help="格式化重试轮数（默认 1）")
    ap.add_argument("--download_hf", action="store_true", help="从 HuggingFace 下载 GAIA（需要 HF_TOKEN 且已在网页同意条款）")
    ap.add_argument("--hf_dir", type=str, default=str(Path(__file__).resolve().parent / "data" / "GAIA"), help="下载目录（--download_hf 生效）")
    ap.add_argument("--subset", type=str, default="2023_all", help="2023_all/2023_level1/2023_level2/2023_level3（--download_hf 生效）")
    ap.add_argument("--split", type=str, default="validation", help="validation/test（--download_hf 生效；test 通常无答案）")
    ap.add_argument("--hf_endpoint", type=str, default=os.environ.get("HF_ENDPOINT") or "https://hf-mirror.com", help="HuggingFace Hub endpoint（默认使用 hf-mirror.com）")
    ap.add_argument("--hf_token", type=str, default=os.environ.get("HF_TOKEN") or os.environ.get("HUGGINGFACE_HUB_TOKEN") or "", help="HF Token（GAIA 为 gated 数据集，需要在 HF 页面同意条款；细粒度 token 需开启 public gated repo access）")
    ap.add_argument("--disable_vector_kb", action="store_true")
    ap.add_argument("--disable_rag", action="store_true")
    ap.add_argument("--disable_tools", action="store_true")
    ap.add_argument(
        "--no_gaia_disable_cdem",
        action="store_true",
        help="GAIA 评测时不自动关闭 CDEM 的向量库/RAG/KB 工具（默认会自动关闭，以避免无关资料干扰）",
    )
    ap.add_argument("--self_test", action="store_true")
    args = ap.parse_args()

    if _env_flag("GAIA_SUPPRESS_WARNINGS", True):
        warnings.filterwarnings("ignore", category=ResourceWarning)
        warnings.filterwarnings("ignore", category=DeprecationWarning, message=r".*swigvarlink.*")

    if args.self_test:
        raise SystemExit(_run_self_test())

    if not bool(args.no_force_gaia_mode):
        os.environ["GAIA_MODE"] = "1"
    gaia_mode = (os.environ.get("GAIA_MODE") or "").strip() in {"1", "true", "on", "yes"}

    if gaia_mode and (not _env_flag("GAIA_KEEP_CUSTOM_SYSTEM_PROMPT", False)):
        os.environ["CDEM_SYSTEM_PROMPT_MODE"] = "generic"

    if gaia_mode and (not bool(args.disable_tools)):
        os.environ["CDEM_ENABLE_TOOLS"] = "1"

    dataset_path = (args.dataset or "").strip()
    if args.download_hf or not dataset_path:
        parquet_path = _download_gaia_hf(
            local_dir=Path(args.hf_dir),
            subset=str(args.subset),
            split=str(args.split),
            endpoint=str(args.hf_endpoint or "").strip() or None,
            token=str(args.hf_token or "").strip() or None,
        )
        dataset_path = str(parquet_path)

    if not dataset_path:
        raise SystemExit("缺少 --dataset（或设置 GAIA_DATASET_PATH）")

    rr = _infer_gaia_repo_root(Path(dataset_path))
    if rr is not None and (os.environ.get("GAIA_ROOT") or "").strip() == "":
        os.environ["GAIA_ROOT"] = str(rr)

    if gaia_mode and (not bool(args.no_gaia_disable_cdem)):
        os.environ["CDEM_ENABLE_VECTOR_KB"] = "0"
        os.environ["CDEM_ENABLE_RAG"] = "0"
        os.environ["CDEM_ENABLE_RAG_FALLBACK"] = "0"
        os.environ["CDEM_ENABLE_KB_TOOL"] = "0"
        if not bool(args.no_force_task_type):
            os.environ["CDEM_TASK_TYPE"] = "qa"

    if args.disable_vector_kb:
        os.environ["CDEM_ENABLE_VECTOR_KB"] = "0"
    if args.disable_rag:
        os.environ["CDEM_ENABLE_RAG"] = "0"
        os.environ["CDEM_ENABLE_RAG_FALLBACK"] = "0"
    if args.disable_tools:
        os.environ["CDEM_ENABLE_TOOLS"] = "0"

    summary = evaluate_gaia(
        dataset_path=dataset_path,
        model_name=args.model,
        output_dir=args.output_dir,
        backend=str(args.backend),
        limit=args.limit,
        offset=args.offset,
        shuffle=bool(args.shuffle),
        seed=int(args.seed),
        level=args.level,
        only_with_file=bool(args.only_with_file),
        inline_attachment=bool(args.force_inline_attachment) or (bool(args.inline_attachment) and (not bool(args.no_inline_attachment))),
        inline_max_chars=int(args.inline_max_chars),
        question_key=args.question_key,
        answer_key=args.answer_key,
        id_key=args.id_key,
        verbose=bool(args.verbose),
        dry_run=bool(args.dry_run),
        timeout_s=args.timeout_s,
        gaia_preamble=(not bool(args.no_gaia_preamble)),
        preamble_text=(
            Path(str(args.gaia_preamble_file)).read_text(encoding="utf-8").strip()
            if str(args.gaia_preamble_file or "").strip()
            else (str(args.gaia_preamble_text or "").strip() or None)
        ),
        format_retry=(not bool(args.no_format_retry)),
        format_retry_rounds=int(args.format_retry_rounds),
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
